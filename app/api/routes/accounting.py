"""
Advanced Accounting + KRA Tax System Routes
Premium-gated endpoints for full financial management.

Endpoints:
  # Entries
  GET    /accounting/entries               – list with filters
  POST   /accounting/entries               – create entry
  PUT    /accounting/entries/{id}          – update entry
  DELETE /accounting/entries/{id}          – delete entry
  POST   /accounting/entries/bulk          – bulk import

  # Sync
  POST   /accounting/sync-payments         – pull confirmed payments → income entries

  # Reports
  GET    /accounting/reports/pnl           – P&L report
  GET    /accounting/reports/cashflow      – monthly cash flow
  GET    /accounting/reports/property-performance  – per-property ROI
  GET    /accounting/reports/expense-breakdown     – expenses by category

  # Tax
  GET    /accounting/tax/summary           – compute tax for a period
  GET    /accounting/tax/constants         – KRA tax constants / reference data
  POST   /accounting/tax/records           – save tax record
  GET    /accounting/tax/records           – list tax records
  PUT    /accounting/tax/records/{id}      – update (mark filed etc.)
  GET    /accounting/tax/withholding       – list withholding entries
  POST   /accounting/tax/withholding       – add withholding entry

  # Export
  GET    /accounting/export/pdf            – PDF P&L + tax summary
  GET    /accounting/export/excel          – Excel workbook (multi-sheet)
  GET    /accounting/export/kra-schedule   – KRA rental income schedule CSV
"""
from __future__ import annotations

import csv
import io
import logging
import uuid as uuid_module
from datetime import datetime, date
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user
from app.models.accounting import (
    AccountingEntry, TaxRecord, WithholdingTaxEntry,
    EntryType, EntryCategory, TaxRecordStatus,
)
from app.models.payment import Subscription, SubscriptionStatus, SubscriptionPlan, Payment, PaymentStatus, PaymentType
from app.models.user import User, UserRole
from app.schemas.accounting import (
    AccountingEntryCreate, AccountingEntryUpdate, AccountingEntryOut,
    BulkImportRequest,
    TaxRecordCreate, TaxRecordUpdate, TaxRecordOut,
    WithholdingEntryCreate, WithholdingEntryOut,
)
from app.services import kra_tax_service

router = APIRouter(tags=["Accounting"])
logger = logging.getLogger(__name__)


# ═══════════════════════ PREMIUM GATE ═══════════════════════

def require_premium(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> User:
    """Allow ADMIN freely; require Professional/Enterprise subscription for others."""
    if current_user.role == UserRole.ADMIN:
        return current_user
    sub = (
        db.query(Subscription)
        .filter(
            Subscription.user_id == current_user.id,
            Subscription.status == SubscriptionStatus.ACTIVE,
            Subscription.plan.in_([SubscriptionPlan.PROFESSIONAL, SubscriptionPlan.ENTERPRISE]),
        )
        .first()
    )
    if not sub:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Advanced Accounting requires a Professional or Enterprise subscription.",
        )
    return current_user


# ═══════════════════════ SERIALISERS ═══════════════════════

def _entry_to_dict(e: AccountingEntry) -> dict:
    def _str(v) -> Optional[str]:
        return str(v) if v is not None else None

    def _dt(v) -> Optional[str]:
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.isoformat()
        if isinstance(v, date):
            return v.strftime("%Y-%m-%d")
        return str(v)

    return {
        "id": str(e.id),
        "owner_id": str(e.owner_id),
        "property_id": _str(e.property_id),
        "unit_id": _str(e.unit_id),
        "tenant_id": _str(e.tenant_id),
        "entry_type": e.entry_type.value if hasattr(e.entry_type, "value") else str(e.entry_type),
        "category": e.category.value if hasattr(e.category, "value") else str(e.category),
        "amount": float(e.amount),
        "description": e.description,
        "reference_number": e.reference_number,
        "entry_date": _dt(e.entry_date),
        "tax_period": e.tax_period,
        "is_reconciled": e.is_reconciled,
        "receipt_url": e.receipt_url,
        "synced_from_payment_id": _str(e.synced_from_payment_id),
        "created_at": _dt(e.created_at),
        "updated_at": _dt(e.updated_at),
    }


def _tax_record_to_dict(t: TaxRecord) -> dict:
    def _dt(v) -> Optional[str]:
        return v.isoformat() if isinstance(v, datetime) else None

    return {
        "id": str(t.id),
        "owner_id": str(t.owner_id),
        "tax_year": t.tax_year,
        "tax_period": t.tax_period,
        "gross_rental_income": float(t.gross_rental_income),
        "allowable_deductions": float(t.allowable_deductions),
        "net_taxable_income": float(t.net_taxable_income),
        "tax_liability": float(t.tax_liability),
        "tax_rate_applied": float(t.tax_rate_applied),
        "landlord_type": t.landlord_type,
        "kra_pin": t.kra_pin,
        "above_threshold": t.above_threshold,
        "status": t.status.value if hasattr(t.status, "value") else str(t.status),
        "filed_at": _dt(t.filed_at),
        "notes": t.notes,
        "created_at": _dt(t.created_at),
        "updated_at": _dt(t.updated_at),
    }


def _wht_to_dict(w: WithholdingTaxEntry) -> dict:
    def _dt(v) -> Optional[str]:
        return v.isoformat() if isinstance(v, datetime) else None

    return {
        "id": str(w.id),
        "owner_id": str(w.owner_id),
        "tenant_id": str(w.tenant_id) if w.tenant_id else None,
        "property_id": str(w.property_id) if w.property_id else None,
        "amount_paid": float(w.amount_paid),
        "withholding_rate": float(w.withholding_rate),
        "withholding_amount": float(w.withholding_amount),
        "period": w.period,
        "certificate_number": w.certificate_number,
        "certificate_url": w.certificate_url,
        "tenant_name": w.tenant_name,
        "tenant_kra_pin": w.tenant_kra_pin,
        "notes": w.notes,
        "created_at": _dt(w.created_at),
    }


# ═══════════════════════ HELPERS ═══════════════════════

def _derive_tax_period(entry_date_str: str) -> str:
    """Derive 'YYYY-MM' from an entry date string."""
    try:
        dt = datetime.strptime(entry_date_str[:10], "%Y-%m-%d")
        return dt.strftime("%Y-%m")
    except Exception:
        return datetime.utcnow().strftime("%Y-%m")


def _try_uuid(v: Optional[str]) -> Optional[uuid_module.UUID]:
    if not v:
        return None
    try:
        return uuid_module.UUID(str(v))
    except Exception:
        return None


def _entry_date_to_dt(entry_date_str: str) -> datetime:
    """Parse entry_date string → datetime."""
    try:
        return datetime.strptime(entry_date_str[:10], "%Y-%m-%d")
    except Exception:
        return datetime.utcnow()


# ═══════════════════════ ENTRY ENDPOINTS ═══════════════════════

@router.get("/entries")
def list_entries(
    property_id: Optional[str] = Query(None),
    entry_type: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),   # "YYYY-MM-DD"
    date_to: Optional[str] = Query(None),     # "YYYY-MM-DD"
    tax_period: Optional[str] = Query(None),  # "YYYY-MM"
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    q = db.query(AccountingEntry).filter(AccountingEntry.owner_id == current_user.id)

    if property_id:
        pid = _try_uuid(property_id)
        if pid:
            q = q.filter(AccountingEntry.property_id == pid)

    if entry_type:
        try:
            q = q.filter(AccountingEntry.entry_type == EntryType(entry_type))
        except ValueError:
            pass

    if category:
        try:
            q = q.filter(AccountingEntry.category == EntryCategory(category))
        except ValueError:
            pass

    if tax_period:
        q = q.filter(AccountingEntry.tax_period == tax_period)

    if date_from:
        try:
            df = datetime.strptime(date_from, "%Y-%m-%d")
            q = q.filter(AccountingEntry.entry_date >= df)
        except ValueError:
            pass

    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            q = q.filter(AccountingEntry.entry_date <= dt)
        except ValueError:
            pass

    total = q.count()
    entries = q.order_by(AccountingEntry.entry_date.desc()).offset(skip).limit(limit).all()

    return {
        "success": True,
        "total": total,
        "entries": [_entry_to_dict(e) for e in entries],
    }


@router.post("/entries")
def create_entry(
    payload: AccountingEntryCreate,
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    tax_period = payload.tax_period or _derive_tax_period(payload.entry_date)

    try:
        entry_type_enum = EntryType(payload.entry_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid entry_type: {payload.entry_type}")

    try:
        category_enum = EntryCategory(payload.category)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid category: {payload.category}")

    entry = AccountingEntry(
        owner_id=current_user.id,
        property_id=_try_uuid(payload.property_id),
        unit_id=_try_uuid(payload.unit_id),
        tenant_id=_try_uuid(payload.tenant_id),
        entry_type=entry_type_enum,
        category=category_enum,
        amount=payload.amount,
        description=payload.description,
        reference_number=payload.reference_number,
        entry_date=_entry_date_to_dt(payload.entry_date),
        tax_period=tax_period,
        is_reconciled=payload.is_reconciled,
        receipt_url=payload.receipt_url,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"success": True, "entry": _entry_to_dict(entry)}


@router.put("/entries/{entry_id}")
def update_entry(
    entry_id: str,
    payload: AccountingEntryUpdate,
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    eid = _try_uuid(entry_id)
    if not eid:
        raise HTTPException(status_code=400, detail="Invalid entry_id")

    entry = (
        db.query(AccountingEntry)
        .filter(AccountingEntry.id == eid, AccountingEntry.owner_id == current_user.id)
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    if payload.entry_type is not None:
        try:
            entry.entry_type = EntryType(payload.entry_type)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid entry_type: {payload.entry_type}")

    if payload.category is not None:
        try:
            entry.category = EntryCategory(payload.category)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid category: {payload.category}")

    if payload.amount is not None:
        entry.amount = payload.amount
    if payload.description is not None:
        entry.description = payload.description
    if payload.reference_number is not None:
        entry.reference_number = payload.reference_number
    if payload.entry_date is not None:
        entry.entry_date = _entry_date_to_dt(payload.entry_date)
        entry.tax_period = payload.tax_period or _derive_tax_period(payload.entry_date)
    if payload.tax_period is not None:
        entry.tax_period = payload.tax_period
    if payload.property_id is not None:
        entry.property_id = _try_uuid(payload.property_id)
    if payload.unit_id is not None:
        entry.unit_id = _try_uuid(payload.unit_id)
    if payload.tenant_id is not None:
        entry.tenant_id = _try_uuid(payload.tenant_id)
    if payload.is_reconciled is not None:
        entry.is_reconciled = payload.is_reconciled
    if payload.receipt_url is not None:
        entry.receipt_url = payload.receipt_url

    entry.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(entry)
    return {"success": True, "entry": _entry_to_dict(entry)}


@router.delete("/entries/{entry_id}")
def delete_entry(
    entry_id: str,
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    eid = _try_uuid(entry_id)
    if not eid:
        raise HTTPException(status_code=400, detail="Invalid entry_id")

    entry = (
        db.query(AccountingEntry)
        .filter(AccountingEntry.id == eid, AccountingEntry.owner_id == current_user.id)
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    db.delete(entry)
    db.commit()
    return {"success": True, "message": "Entry deleted"}


@router.post("/entries/bulk")
def bulk_import_entries(
    payload: BulkImportRequest,
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    created = []
    errors = []

    for idx, row in enumerate(payload.entries):
        try:
            entry_type_enum = EntryType(row.entry_type)
            category_enum = EntryCategory(row.category)
            tax_period = _derive_tax_period(row.entry_date)
            entry = AccountingEntry(
                owner_id=current_user.id,
                property_id=_try_uuid(row.property_id),
                tenant_id=_try_uuid(row.tenant_id),
                entry_type=entry_type_enum,
                category=category_enum,
                amount=row.amount,
                description=row.description,
                reference_number=row.reference_number,
                entry_date=_entry_date_to_dt(row.entry_date),
                tax_period=tax_period,
            )
            db.add(entry)
            created.append(idx)
        except Exception as exc:
            errors.append({"row": idx, "error": str(exc)})

    db.commit()
    return {
        "success": True,
        "created": len(created),
        "errors": errors,
    }


# ═══════════════════════ SYNC PAYMENTS ═══════════════════════

@router.post("/sync-payments")
def sync_payments(
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    """
    Pull all COMPLETED rent/deposit payments for this owner
    and create corresponding income accounting entries.
    Idempotent — already-synced payments are skipped.
    """
    # Find all payments for this owner's tenants
    completed_payments = (
        db.query(Payment)
        .filter(
            Payment.user_id == current_user.id,
            Payment.status == PaymentStatus.COMPLETED,
            Payment.payment_type.in_([
                PaymentType.RENT,
                PaymentType.DEPOSIT,
                PaymentType.LATE_FEE if hasattr(PaymentType, "LATE_FEE") else PaymentType.PENALTY,
            ]),
        )
        .all()
    )

    synced = 0
    skipped = 0

    for pmt in completed_payments:
        # Check if already synced
        existing = (
            db.query(AccountingEntry)
            .filter(AccountingEntry.synced_from_payment_id == pmt.id)
            .first()
        )
        if existing:
            skipped += 1
            continue

        # Map payment type → accounting category
        if pmt.payment_type in (PaymentType.RENT,):
            cat = EntryCategory.RENTAL_INCOME
        elif pmt.payment_type == PaymentType.DEPOSIT:
            cat = EntryCategory.DEPOSIT_RECEIVED
        else:
            cat = EntryCategory.OTHER_INCOME

        entry_date = pmt.paid_at or pmt.payment_date or pmt.created_at
        tax_period = entry_date.strftime("%Y-%m") if entry_date else datetime.utcnow().strftime("%Y-%m")

        entry = AccountingEntry(
            owner_id=current_user.id,
            tenant_id=pmt.tenant_id,
            entry_type=EntryType.INCOME,
            category=cat,
            amount=float(pmt.amount),
            description=pmt.description or f"Payment ref: {pmt.reference}",
            reference_number=pmt.reference,
            entry_date=entry_date or datetime.utcnow(),
            tax_period=tax_period,
            is_reconciled=True,
            synced_from_payment_id=pmt.id,
        )
        db.add(entry)
        synced += 1

    db.commit()
    return {
        "success": True,
        "synced": synced,
        "skipped_already_synced": skipped,
        "message": f"Synced {synced} payments ({skipped} already existed)",
    }


# ═══════════════════════ REPORT ENDPOINTS ═══════════════════════

def _get_period_bounds(year: int, month: Optional[int], period: str):
    """Return (date_from, date_to, label) for the given period parameters."""
    if period == "monthly" and month:
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        date_from = datetime(year, month, 1)
        date_to = datetime(year, month, last_day, 23, 59, 59)
        label = f"{year}-{month:02d}"
    elif period == "quarterly" and month:
        quarter = (month - 1) // 3 + 1
        q_start_month = (quarter - 1) * 3 + 1
        q_end_month = q_start_month + 2
        import calendar
        last_day = calendar.monthrange(year, q_end_month)[1]
        date_from = datetime(year, q_start_month, 1)
        date_to = datetime(year, q_end_month, last_day, 23, 59, 59)
        label = f"{year}-Q{quarter}"
    else:
        date_from = datetime(year, 1, 1)
        date_to = datetime(year, 12, 31, 23, 59, 59)
        label = str(year)
    return date_from, date_to, label


@router.get("/reports/pnl")
def pnl_report(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    period: str = Query("monthly"),      # monthly | quarterly | annual
    property_id: Optional[str] = Query(None),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    date_from, date_to, label = _get_period_bounds(year, month, period)

    q = db.query(AccountingEntry).filter(
        AccountingEntry.owner_id == current_user.id,
        AccountingEntry.entry_date >= date_from,
        AccountingEntry.entry_date <= date_to,
    )
    if property_id:
        pid = _try_uuid(property_id)
        if pid:
            q = q.filter(AccountingEntry.property_id == pid)

    entries = q.all()

    income_breakdown: dict = {}
    expense_breakdown: dict = {}
    gross_income = 0.0
    total_expenses = 0.0

    for e in entries:
        cat = e.category.value if hasattr(e.category, "value") else str(e.category)
        amt = float(e.amount)
        if e.entry_type == EntryType.INCOME or str(e.entry_type) == "income":
            gross_income += amt
            income_breakdown[cat] = income_breakdown.get(cat, 0.0) + amt
        else:
            total_expenses += amt
            expense_breakdown[cat] = expense_breakdown.get(cat, 0.0) + amt

    net_profit = gross_income - total_expenses
    net_margin = (net_profit / gross_income * 100) if gross_income > 0 else 0.0

    return {
        "success": True,
        "report": {
            "period": label,
            "period_type": period,
            "property_id": property_id,
            "gross_income": round(gross_income, 2),
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(net_profit, 2),
            "net_margin_pct": round(net_margin, 2),
            "income_breakdown": {k: round(v, 2) for k, v in income_breakdown.items()},
            "expense_breakdown": {k: round(v, 2) for k, v in expense_breakdown.items()},
            "entry_count": len(entries),
        },
    }


@router.get("/reports/cashflow")
def cashflow_report(
    year: int = Query(...),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    """Monthly cash flow: income vs expenses for every month in the year."""
    entries = (
        db.query(AccountingEntry)
        .filter(
            AccountingEntry.owner_id == current_user.id,
            AccountingEntry.entry_date >= datetime(year, 1, 1),
            AccountingEntry.entry_date <= datetime(year, 12, 31, 23, 59, 59),
        )
        .all()
    )

    months: dict = {f"{year}-{m:02d}": {"income": 0.0, "expenses": 0.0} for m in range(1, 13)}

    for e in entries:
        period = e.entry_date.strftime("%Y-%m") if e.entry_date else None
        if period and period in months:
            if e.entry_type == EntryType.INCOME or str(e.entry_type) == "income":
                months[period]["income"] += float(e.amount)
            else:
                months[period]["expenses"] += float(e.amount)

    cashflow = [
        {
            "month": period,
            "income": round(data["income"], 2),
            "expenses": round(data["expenses"], 2),
            "net": round(data["income"] - data["expenses"], 2),
        }
        for period, data in sorted(months.items())
    ]

    return {"success": True, "year": year, "cashflow": cashflow}


@router.get("/reports/property-performance")
def property_performance_report(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    """Per-property ROI: gross income, expenses, net profit, occupancy (from units)."""
    date_from, date_to, label = _get_period_bounds(year, month, "monthly" if month else "annual")

    entries = (
        db.query(AccountingEntry)
        .filter(
            AccountingEntry.owner_id == current_user.id,
            AccountingEntry.entry_date >= date_from,
            AccountingEntry.entry_date <= date_to,
            AccountingEntry.property_id.isnot(None),
        )
        .all()
    )

    # Group by property
    by_prop: dict = {}
    for e in entries:
        pid = str(e.property_id)
        if pid not in by_prop:
            by_prop[pid] = {"income": 0.0, "expenses": 0.0}
        if e.entry_type == EntryType.INCOME or str(e.entry_type) == "income":
            by_prop[pid]["income"] += float(e.amount)
        else:
            by_prop[pid]["expenses"] += float(e.amount)

    # Enrich with property names and unit occupancy
    try:
        from app.models.property import Property, Unit
        results = []
        for pid, data in by_prop.items():
            prop = db.query(Property).filter(Property.id == _try_uuid(pid)).first()
            prop_name = prop.name if prop else f"Property {pid[:8]}"

            # Occupancy rate
            try:
                total_units = db.query(func.count(Unit.id)).filter(Unit.property_id == _try_uuid(pid)).scalar() or 0
                occupied_units = (
                    db.query(func.count(Unit.id))
                    .filter(Unit.property_id == _try_uuid(pid), Unit.status == "occupied")
                    .scalar() or 0
                )
                occupancy = (occupied_units / total_units * 100) if total_units > 0 else 0.0
            except Exception:
                occupancy = 0.0

            net = data["income"] - data["expenses"]
            results.append({
                "property_id": pid,
                "property_name": prop_name,
                "gross_income": round(data["income"], 2),
                "total_expenses": round(data["expenses"], 2),
                "net_profit": round(net, 2),
                "occupancy_rate": round(occupancy, 1),
                "gross_yield_pct": None,  # Would need property purchase price
            })
    except Exception as exc:
        logger.warning(f"Property performance enrichment failed: {exc}")
        results = [
            {
                "property_id": pid,
                "property_name": f"Property {pid[:8]}",
                "gross_income": round(data["income"], 2),
                "total_expenses": round(data["expenses"], 2),
                "net_profit": round(data["income"] - data["expenses"], 2),
                "occupancy_rate": 0.0,
                "gross_yield_pct": None,
            }
            for pid, data in by_prop.items()
        ]

    return {"success": True, "period": label, "properties": results}


@router.get("/reports/expense-breakdown")
def expense_breakdown_report(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    property_id: Optional[str] = Query(None),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    date_from, date_to, label = _get_period_bounds(year, month, "monthly" if month else "annual")

    q = db.query(AccountingEntry).filter(
        AccountingEntry.owner_id == current_user.id,
        AccountingEntry.entry_date >= date_from,
        AccountingEntry.entry_date <= date_to,
    )
    if property_id:
        pid = _try_uuid(property_id)
        if pid:
            q = q.filter(AccountingEntry.property_id == pid)

    entries = q.all()

    total_income = sum(float(e.amount) for e in entries if e.entry_type == EntryType.INCOME or str(e.entry_type) == "income")
    expense_by_cat: dict = {}
    for e in entries:
        if e.entry_type == EntryType.EXPENSE or str(e.entry_type) == "expense":
            cat = e.category.value if hasattr(e.category, "value") else str(e.category)
            expense_by_cat[cat] = expense_by_cat.get(cat, 0.0) + float(e.amount)

    total_expenses = sum(expense_by_cat.values())
    breakdown = [
        {
            "category": cat,
            "amount": round(amt, 2),
            "pct_of_expenses": round(amt / total_expenses * 100, 1) if total_expenses > 0 else 0,
            "pct_of_income": round(amt / total_income * 100, 1) if total_income > 0 else 0,
        }
        for cat, amt in sorted(expense_by_cat.items(), key=lambda x: -x[1])
    ]

    return {
        "success": True,
        "period": label,
        "total_income": round(total_income, 2),
        "total_expenses": round(total_expenses, 2),
        "breakdown": breakdown,
    }


# ═══════════════════════ TAX ENDPOINTS ═══════════════════════

@router.get("/tax/constants")
def get_tax_constants(current_user: User = Depends(get_current_user)):
    """Return KRA tax constants for frontend reference (no premium gate)."""
    return {"success": True, "constants": kra_tax_service.get_tax_constants()}


@router.get("/tax/summary")
def get_tax_summary(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    period_type: str = Query("monthly"),   # monthly | annual
    landlord_type: str = Query("resident_individual"),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    """
    Compute KRA tax liability for the given period.
    Pulls actual income and allowable-deduction entries from the ledger.
    """
    date_from, date_to, label = _get_period_bounds(year, month, period_type)

    entries = (
        db.query(AccountingEntry)
        .filter(
            AccountingEntry.owner_id == current_user.id,
            AccountingEntry.entry_date >= date_from,
            AccountingEntry.entry_date <= date_to,
        )
        .all()
    )

    # Sum gross income
    gross_income = sum(
        float(e.amount) for e in entries
        if e.entry_type == EntryType.INCOME or str(e.entry_type) == "income"
    )

    # Sum allowable deductions (only KRA-recognised expense categories)
    allowable_cats = set(kra_tax_service.get_allowable_categories())
    total_deductions = sum(
        float(e.amount) for e in entries
        if (e.entry_type == EntryType.EXPENSE or str(e.entry_type) == "expense")
        and (e.category.value if hasattr(e.category, "value") else str(e.category)) in allowable_cats
    )

    # Compute annual gross for threshold check (for monthly, annualise)
    if period_type == "monthly":
        annualised = gross_income * 12
        result = kra_tax_service.compute_monthly_tax(
            gross_monthly_rent=gross_income,
            landlord_type=landlord_type,
            total_allowable_deductions=total_deductions,
            annualised_gross=annualised,
        )
    else:
        result = kra_tax_service.compute_annual_tax(
            gross_annual_rent=gross_income,
            landlord_type=landlord_type,
            total_allowable_deductions=total_deductions,
        )

    result["period"] = label
    result["entry_count"] = len(entries)

    return {"success": True, "tax_summary": result}


@router.post("/tax/records")
def create_tax_record(
    payload: TaxRecordCreate,
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    record = TaxRecord(
        owner_id=current_user.id,
        tax_year=payload.tax_year,
        tax_period=payload.tax_period,
        gross_rental_income=payload.gross_rental_income,
        allowable_deductions=payload.allowable_deductions,
        net_taxable_income=payload.net_taxable_income,
        tax_liability=payload.tax_liability,
        tax_rate_applied=payload.tax_rate_applied,
        landlord_type=payload.landlord_type,
        kra_pin=payload.kra_pin,
        above_threshold=payload.above_threshold,
        status=TaxRecordStatus(payload.status),
        notes=payload.notes,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return {"success": True, "record": _tax_record_to_dict(record)}


@router.get("/tax/records")
def list_tax_records(
    year: Optional[int] = Query(None),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    q = db.query(TaxRecord).filter(TaxRecord.owner_id == current_user.id)
    if year:
        q = q.filter(TaxRecord.tax_year == year)
    records = q.order_by(TaxRecord.tax_year.desc(), TaxRecord.created_at.desc()).all()
    return {"success": True, "records": [_tax_record_to_dict(r) for r in records]}


@router.put("/tax/records/{record_id}")
def update_tax_record(
    record_id: str,
    payload: TaxRecordUpdate,
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    rid = _try_uuid(record_id)
    if not rid:
        raise HTTPException(status_code=400, detail="Invalid record_id")

    record = (
        db.query(TaxRecord)
        .filter(TaxRecord.id == rid, TaxRecord.owner_id == current_user.id)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="Tax record not found")

    if payload.status is not None:
        try:
            record.status = TaxRecordStatus(payload.status)
            if payload.status == "filed" and record.filed_at is None:
                record.filed_at = datetime.utcnow()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {payload.status}")
    if payload.kra_pin is not None:
        record.kra_pin = payload.kra_pin
    if payload.notes is not None:
        record.notes = payload.notes
    if payload.filed_at is not None:
        try:
            record.filed_at = datetime.fromisoformat(payload.filed_at)
        except ValueError:
            pass

    record.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(record)
    return {"success": True, "record": _tax_record_to_dict(record)}


@router.get("/tax/withholding")
def list_withholding(
    period: Optional[str] = Query(None),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    q = db.query(WithholdingTaxEntry).filter(WithholdingTaxEntry.owner_id == current_user.id)
    if period:
        q = q.filter(WithholdingTaxEntry.period == period)
    entries = q.order_by(WithholdingTaxEntry.period.desc()).all()
    return {"success": True, "entries": [_wht_to_dict(e) for e in entries]}


@router.post("/tax/withholding")
def create_withholding_entry(
    payload: WithholdingEntryCreate,
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    wht_amount = round(payload.amount_paid * payload.withholding_rate / 100, 2)
    entry = WithholdingTaxEntry(
        owner_id=current_user.id,
        tenant_id=_try_uuid(payload.tenant_id),
        property_id=_try_uuid(payload.property_id),
        amount_paid=payload.amount_paid,
        withholding_rate=payload.withholding_rate,
        withholding_amount=wht_amount,
        period=payload.period,
        certificate_number=payload.certificate_number,
        certificate_url=payload.certificate_url,
        tenant_name=payload.tenant_name,
        tenant_kra_pin=payload.tenant_kra_pin,
        notes=payload.notes,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"success": True, "entry": _wht_to_dict(entry)}


# ═══════════════════════ EXPORT ENDPOINTS ═══════════════════════

@router.get("/export/pdf")
def export_pdf(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    period_type: str = Query("monthly"),
    landlord_type: str = Query("resident_individual"),
    property_id: Optional[str] = Query(None),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    """Generate a PDF P&L + Tax Summary report using reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    date_from, date_to, label = _get_period_bounds(year, month, period_type)

    # Fetch data
    q = db.query(AccountingEntry).filter(
        AccountingEntry.owner_id == current_user.id,
        AccountingEntry.entry_date >= date_from,
        AccountingEntry.entry_date <= date_to,
    )
    if property_id:
        pid = _try_uuid(property_id)
        if pid:
            q = q.filter(AccountingEntry.property_id == pid)
    entries = q.all()

    gross_income = 0.0
    total_expenses = 0.0
    income_breakdown: dict = {}
    expense_breakdown: dict = {}

    for e in entries:
        cat = e.category.value if hasattr(e.category, "value") else str(e.category)
        amt = float(e.amount)
        if e.entry_type == EntryType.INCOME or str(e.entry_type) == "income":
            gross_income += amt
            income_breakdown[cat] = income_breakdown.get(cat, 0.0) + amt
        else:
            total_expenses += amt
            expense_breakdown[cat] = expense_breakdown.get(cat, 0.0) + amt

    net_profit = gross_income - total_expenses

    # Tax computation
    allowable_cats = set(kra_tax_service.get_allowable_categories())
    total_deductions = sum(v for k, v in expense_breakdown.items() if k in allowable_cats)
    if period_type == "monthly":
        tax_result = kra_tax_service.compute_monthly_tax(
            gross_income, landlord_type, total_deductions, gross_income * 12
        )
    else:
        tax_result = kra_tax_service.compute_annual_tax(gross_income, landlord_type, total_deductions)

    # Build PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    BLUE = colors.HexColor("#1a56db")
    LIGHT_GREY = colors.HexColor("#f3f4f6")

    title_style = ParagraphStyle("title", parent=styles["Heading1"], textColor=BLUE, fontSize=18, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], textColor=colors.grey, fontSize=10)
    section_style = ParagraphStyle("section", parent=styles["Heading2"], textColor=BLUE, fontSize=13, spaceBefore=14, spaceAfter=4)

    story = [
        Paragraph("PROPERTECH — Financial Report", title_style),
        Paragraph(f"Period: {label}  |  Landlord Type: {kra_tax_service.LANDLORD_TYPES.get(landlord_type, landlord_type)}", sub_style),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", sub_style),
        Spacer(1, 0.5 * cm),
    ]

    def _kes(v: float) -> str:
        return f"KES {v:,.2f}"

    def _table(data, col_widths=None):
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d1d5db")),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    # P&L Summary
    story.append(Paragraph("Profit & Loss Summary", section_style))
    pnl_data = [
        ["Item", "Amount (KES)"],
        ["Gross Rental Income", _kes(gross_income)],
        ["Total Expenses", _kes(total_expenses)],
        ["Net Profit", _kes(net_profit)],
        ["Net Margin", f"{(net_profit/gross_income*100):.1f}%" if gross_income > 0 else "0%"],
    ]
    story.append(_table(pnl_data, [10 * cm, 7 * cm]))
    story.append(Spacer(1, 0.4 * cm))

    # Income Breakdown
    if income_breakdown:
        story.append(Paragraph("Income Breakdown", section_style))
        income_data = [["Category", "Amount (KES)"]] + [
            [cat.replace("_", " ").title(), _kes(amt)]
            for cat, amt in sorted(income_breakdown.items(), key=lambda x: -x[1])
        ]
        story.append(_table(income_data, [10 * cm, 7 * cm]))
        story.append(Spacer(1, 0.4 * cm))

    # Expense Breakdown
    if expense_breakdown:
        story.append(Paragraph("Expense Breakdown", section_style))
        expense_data = [["Category", "Amount (KES)"]] + [
            [cat.replace("_", " ").title(), _kes(amt)]
            for cat, amt in sorted(expense_breakdown.items(), key=lambda x: -x[1])
        ]
        story.append(_table(expense_data, [10 * cm, 7 * cm]))
        story.append(Spacer(1, 0.4 * cm))

    # KRA Tax Summary
    story.append(Paragraph("KRA Tax Computation", section_style))
    tax_data = [
        ["Item", "Value"],
        ["Gross Rental Income", _kes(tax_result["gross_rental_income"])],
        ["Allowable Deductions", _kes(tax_result["allowable_deductions"])],
        ["Net Taxable Income", _kes(tax_result["net_taxable_income"])],
        ["Tax Rate Applied", f"{tax_result['tax_rate_applied'] * 100:.2f}%"],
        ["Tax Liability", _kes(tax_result["tax_liability"])],
        ["Calculation Method", tax_result["calculation_method"][:80]],
        ["Above KES 15M Threshold", "YES" if tax_result["above_mri_threshold"] else "NO"],
    ]
    story.append(_table(tax_data, [10 * cm, 7 * cm]))

    doc.build(story)
    buf.seek(0)

    filename = f"propertech_financial_report_{label.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/excel")
def export_excel(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    period_type: str = Query("monthly"),
    landlord_type: str = Query("resident_individual"),
    property_id: Optional[str] = Query(None),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    """Generate a multi-sheet Excel workbook: Income, Expenses, P&L, Tax Summary, Withholding."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    date_from, date_to, label = _get_period_bounds(year, month, period_type)

    q = db.query(AccountingEntry).filter(
        AccountingEntry.owner_id == current_user.id,
        AccountingEntry.entry_date >= date_from,
        AccountingEntry.entry_date <= date_to,
    )
    if property_id:
        pid = _try_uuid(property_id)
        if pid:
            q = q.filter(AccountingEntry.property_id == pid)
    entries = q.all()

    income_entries = [e for e in entries if e.entry_type == EntryType.INCOME or str(e.entry_type) == "income"]
    expense_entries = [e for e in entries if e.entry_type == EntryType.EXPENSE or str(e.entry_type) == "expense"]

    # Withholding entries for the period
    wht_entries = (
        db.query(WithholdingTaxEntry)
        .filter(
            WithholdingTaxEntry.owner_id == current_user.id,
            WithholdingTaxEntry.period.between(
                date_from.strftime("%Y-%m"), date_to.strftime("%Y-%m")
            ),
        )
        .all()
    )

    wb = openpyxl.Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1a56db")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TOTAL_FILL = PatternFill("solid", fgColor="e5edff")
    TOTAL_FONT = Font(bold=True, size=10)
    KES_FMT = '#,##0.00'

    def _header_row(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 1: Income ──
    ws_income = wb.active
    ws_income.title = "Income"
    _header_row(ws_income, ["Date", "Category", "Description", "Reference", "Property", "Amount (KES)", "Reconciled"])
    for r, e in enumerate(income_entries, 2):
        ws_income.cell(r, 1, e.entry_date.strftime("%Y-%m-%d") if e.entry_date else "")
        ws_income.cell(r, 2, (e.category.value if hasattr(e.category, "value") else str(e.category)).replace("_", " ").title())
        ws_income.cell(r, 3, e.description or "")
        ws_income.cell(r, 4, e.reference_number or "")
        ws_income.cell(r, 5, str(e.property_id) if e.property_id else "")
        amt_cell = ws_income.cell(r, 6, float(e.amount))
        amt_cell.number_format = KES_FMT
        ws_income.cell(r, 7, "Yes" if e.is_reconciled else "No")
    # Totals row
    tr = len(income_entries) + 2
    ws_income.cell(tr, 5, "TOTAL").font = TOTAL_FONT
    total_cell = ws_income.cell(tr, 6, f"=SUM(F2:F{tr-1})")
    total_cell.number_format = KES_FMT
    total_cell.font = TOTAL_FONT
    total_cell.fill = TOTAL_FILL
    _auto_width(ws_income)

    # ── Sheet 2: Expenses ──
    ws_exp = wb.create_sheet("Expenses")
    _header_row(ws_exp, ["Date", "Category", "Description", "Reference", "Property", "Amount (KES)", "Reconciled"])
    for r, e in enumerate(expense_entries, 2):
        ws_exp.cell(r, 1, e.entry_date.strftime("%Y-%m-%d") if e.entry_date else "")
        ws_exp.cell(r, 2, (e.category.value if hasattr(e.category, "value") else str(e.category)).replace("_", " ").title())
        ws_exp.cell(r, 3, e.description or "")
        ws_exp.cell(r, 4, e.reference_number or "")
        ws_exp.cell(r, 5, str(e.property_id) if e.property_id else "")
        amt_cell = ws_exp.cell(r, 6, float(e.amount))
        amt_cell.number_format = KES_FMT
        ws_exp.cell(r, 7, "Yes" if e.is_reconciled else "No")
    tr2 = len(expense_entries) + 2
    ws_exp.cell(tr2, 5, "TOTAL").font = TOTAL_FONT
    total_exp_cell = ws_exp.cell(tr2, 6, f"=SUM(F2:F{tr2-1})")
    total_exp_cell.number_format = KES_FMT
    total_exp_cell.font = TOTAL_FONT
    total_exp_cell.fill = TOTAL_FILL
    _auto_width(ws_exp)

    # ── Sheet 3: P&L ──
    ws_pnl = wb.create_sheet("P&L")
    gross = sum(float(e.amount) for e in income_entries)
    expenses = sum(float(e.amount) for e in expense_entries)
    net = gross - expenses
    pnl_rows = [
        ("Period", label),
        ("Gross Rental Income", gross),
        ("Total Expenses", expenses),
        ("Net Profit", net),
        ("Net Margin %", f"{(net/gross*100):.2f}%" if gross > 0 else "0%"),
    ]
    for r, (label_txt, val) in enumerate(pnl_rows, 1):
        ws_pnl.cell(r, 1, label_txt).font = Font(bold=True)
        cell = ws_pnl.cell(r, 2, val)
        if isinstance(val, float):
            cell.number_format = KES_FMT
    _auto_width(ws_pnl)

    # ── Sheet 4: Tax Summary ──
    ws_tax = wb.create_sheet("Tax Summary")
    allowable_cats = set(kra_tax_service.get_allowable_categories())
    exp_by_cat = {}
    for e in expense_entries:
        cat = e.category.value if hasattr(e.category, "value") else str(e.category)
        exp_by_cat[cat] = exp_by_cat.get(cat, 0.0) + float(e.amount)
    deductions = sum(v for k, v in exp_by_cat.items() if k in allowable_cats)

    if period_type == "monthly":
        tax_res = kra_tax_service.compute_monthly_tax(gross, landlord_type, deductions, gross * 12)
    else:
        tax_res = kra_tax_service.compute_annual_tax(gross, landlord_type, deductions)

    tax_rows = [
        ("KRA Tax Computation", ""),
        ("Period", label),
        ("Landlord Type", kra_tax_service.LANDLORD_TYPES.get(landlord_type, landlord_type)),
        ("Gross Rental Income (KES)", tax_res["gross_rental_income"]),
        ("Allowable Deductions (KES)", tax_res["allowable_deductions"]),
        ("Net Taxable Income (KES)", tax_res["net_taxable_income"]),
        ("Tax Rate Applied", f"{tax_res['tax_rate_applied'] * 100:.2f}%"),
        ("Tax Liability (KES)", tax_res["tax_liability"]),
        ("Above KES 15M Threshold", "YES" if tax_res["above_mri_threshold"] else "NO"),
        ("Calculation Method", tax_res["calculation_method"]),
        ("", ""),
        ("DISCLAIMER", "For KRA iTax filing reference only. Confirm with a licensed tax advisor."),
    ]
    for r, (k, v) in enumerate(tax_rows, 1):
        ws_tax.cell(r, 1, k).font = Font(bold=True)
        cell = ws_tax.cell(r, 2, v)
        if isinstance(v, float):
            cell.number_format = KES_FMT
    _auto_width(ws_tax)

    # ── Sheet 5: Withholding ──
    ws_wht = wb.create_sheet("Withholding Tax")
    _header_row(ws_wht, ["Period", "Tenant", "Tenant KRA PIN", "Gross Rent", "WHT Rate %", "WHT Amount", "Net Received", "Certificate #"])
    for r, w in enumerate(wht_entries, 2):
        ws_wht.cell(r, 1, w.period)
        ws_wht.cell(r, 2, w.tenant_name or "")
        ws_wht.cell(r, 3, w.tenant_kra_pin or "")
        ws_wht.cell(r, 4, float(w.amount_paid + w.withholding_amount)).number_format = KES_FMT
        ws_wht.cell(r, 5, float(w.withholding_rate))
        ws_wht.cell(r, 6, float(w.withholding_amount)).number_format = KES_FMT
        ws_wht.cell(r, 7, float(w.amount_paid)).number_format = KES_FMT
        ws_wht.cell(r, 8, w.certificate_number or "")
    _auto_width(ws_wht)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"propertech_accounts_{label.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/kra-schedule")
def export_kra_schedule(
    year: int = Query(...),
    current_user: User = Depends(require_premium),
    db: Session = Depends(get_db),
):
    """
    Generate KRA rental income schedule in iTax annual return format.
    Columns: Property Address | Annual Rent Received | Allowable Deductions | Net Income
    """
    try:
        from app.models.property import Property
    except ImportError:
        Property = None

    date_from = datetime(year, 1, 1)
    date_to = datetime(year, 12, 31, 23, 59, 59)

    entries = (
        db.query(AccountingEntry)
        .filter(
            AccountingEntry.owner_id == current_user.id,
            AccountingEntry.entry_date >= date_from,
            AccountingEntry.entry_date <= date_to,
        )
        .all()
    )

    allowable_cats = set(kra_tax_service.get_allowable_categories())

    # Group by property
    by_prop: dict = {}
    for e in entries:
        pid = str(e.property_id) if e.property_id else "unspecified"
        if pid not in by_prop:
            by_prop[pid] = {"income": 0.0, "deductions": 0.0}
        cat = e.category.value if hasattr(e.category, "value") else str(e.category)
        if e.entry_type == EntryType.INCOME or str(e.entry_type) == "income":
            by_prop[pid]["income"] += float(e.amount)
        elif cat in allowable_cats:
            by_prop[pid]["deductions"] += float(e.amount)

    # Build CSV
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        f"KRA Rental Income Schedule — Tax Year {year}",
        "", "", "", "",
    ])
    writer.writerow([
        "Property Address / Description",
        "Annual Rent Received (KES)",
        "Allowable Deductions (KES)",
        "Net Rental Income (KES)",
        "Notes",
    ])

    totals = {"income": 0.0, "deductions": 0.0, "net": 0.0}
    for pid, data in by_prop.items():
        net = data["income"] - data["deductions"]
        # Try to fetch property address
        address = "—"
        if Property and pid != "unspecified":
            prop = db.query(Property).filter(Property.id == _try_uuid(pid)).first()
            if prop:
                address = f"{prop.name} — {prop.address or ''}"

        writer.writerow([
            address,
            f"{data['income']:,.2f}",
            f"{data['deductions']:,.2f}",
            f"{net:,.2f}",
            "",
        ])
        totals["income"] += data["income"]
        totals["deductions"] += data["deductions"]
        totals["net"] += net

    writer.writerow([])
    writer.writerow([
        "TOTAL",
        f"{totals['income']:,.2f}",
        f"{totals['deductions']:,.2f}",
        f"{totals['net']:,.2f}",
        "",
    ])
    writer.writerow([])
    writer.writerow(["Generated by PROPERTECH", f"Date: {datetime.utcnow().strftime('%Y-%m-%d')}", "", "", "For iTax filing reference only"])

    buf.seek(0)
    filename = f"kra_rental_income_schedule_{year}.csv"
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),  # utf-8-sig for Excel CSV compatibility
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
